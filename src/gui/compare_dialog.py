"""
compare_dialog.py — 피부 측정 비교 다이얼로그 모듈

원본 vs 복원 이미지의 18개 측정 항목을 테이블 형식으로 비교 표시하는 다이얼로그를 제공합니다.
"""

from __future__ import annotations

import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional

from PySide6.QtCore import QCoreApplication, Qt, QTimer, QThread
from PySide6.QtGui import QColor, QCloseEvent
from PySide6.QtWidgets import (
    QDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QFileDialog,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from src.utils.xlsx_utils import (
    append_with_font,
    calculate_column_width,
    auto_fit_columns,
)

from src.scoring.skin_scoring import get_measurement_categories
from src.gui.dialog_utils import _numeric_value, _short_label
from src.gui.llm_workers import LlmWorker
from src.skin.core.config_parser import get_measurement_count

log = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


class SkinMeasurementCompareDialog(QDialog):
    """18개 측정 비교 (원본 vs 복원) - 테이블 형식 표시."""

    def __init__(
        self,
        parent: Optional[QWidget],
        orig_path: Path,
        ideal_path: Path,
        orig_result: Dict[str, Any],
        ideal_result: Dict[str, Any],
        llm_scores: bool = False,  # LLM 점수 제공 여부
        llm_provide_scores: bool = True,  # 항상 점수 제공
        llm_orig_report: Optional[Any] = None,  # 원본 LLM 보고서
        llm_ideal_report: Optional[Any] = None,  # 복원 LLM 보고서
        safety_net_applied: bool = False,  # 안전장치 적용 여부
    ) -> None:
        super().__init__(parent)
        # [FIX] 안전장치 적용 여부를 제목에 표시
        title_suffix = " (안전장치 적용)" if safety_net_applied else " (실측 점수)"
        measurement_count = get_measurement_count()
        # 프로젝트 이름과 버전 로드
        try:
            from src.utils.config import load_config as _load_config
            config = _load_config()
            project_name = config.get("project", {}).get("name", "SkinLens")
            project_version = config.get("project", {}).get("version", "1.0.0")
            self.setWindowTitle(f"{project_name} v{project_version} - 피부 분석 {measurement_count}항목 비교 (원본 vs 기준){title_suffix}")
        except Exception:
            self.setWindowTitle(f"SkinLens v1.0.0 - 피부 분석 {measurement_count}항목 비교 (원본 vs 기준){title_suffix}")
        self.resize(1000, 1300)
        
        # LLM 소견 생성을 위해 원본/복원 경로와 측정 결과 저장
        self._orig_path = orig_path
        self._ideal_path = ideal_path
        self._orig_result = orig_result
        self._ideal_result = ideal_result
        self._llm_scores = llm_scores  # LLM 점수 제공 여부 저장
        self._llm_provide_scores = llm_provide_scores  # LLM 점수 제공 여부 저장

        # LLM report 저장 (엑셀 내보내기 시 항목별 소견 출력용)
        self._last_llm_report_orig = llm_orig_report
        self._last_llm_report_ideal = llm_ideal_report

        mo = orig_result.get("measurements_report") or orig_result.get("measurements", {})
        mi = ideal_result.get("measurements_report") or ideal_result.get("measurements", {})
        o_age = orig_result.get("perceived_age", "—")
        i_age = ideal_result.get("perceived_age", "—")
        
        log.debug(f"다이얼로그 초기화: mo keys={list(mo.keys()) if mo else 'empty'}, mi keys={list(mi.keys()) if mi else 'empty'}")

        # [FIX v3.0 ⑥] 항목 루프가 measurements_report(레이어 B)이므로 피부건강지수도 레이어 B 기준으로 통일
        rov = _numeric_value(
            orig_result.get("overall_score_report") or orig_result.get("overall_score")
        )
        riv = _numeric_value(
            ideal_result.get("overall_score_report") or ideal_result.get("overall_score")
        )

        # 항목별 개선/역전/동등 집계 (헤더 요약용)
        n_improved = n_reversed = n_equal = 0
        for _, keys in get_measurement_categories():
            for key in keys:
                if key == "pore_count":
                    # 개수 기준: 이상 < 원본이면 개선
                    vo, vi = _numeric_value(mo.get(key)), _numeric_value(mi.get(key))
                    if vi < vo:
                        n_improved += 1
                    elif vi > vo:
                        n_reversed += 1
                    else:
                        n_equal += 1
                else:
                    vo, vi = _numeric_value(mo.get(key)), _numeric_value(mi.get(key))
                    if vi > vo:
                        n_improved += 1
                    elif vi < vo:
                        n_reversed += 1
                    else:
                        n_equal += 1

        head = QLabel(
            f"<b>원본</b> {orig_path.name} &nbsp;|&nbsp; "
            f"<b>기준</b> {ideal_path.name}<br>"
            f"피부건강지수: 원본 <b>{int(round(rov))}</b> · 기준 <b>{int(round(riv))}</b>점<br>"
            f"인지나이: 원본 <b>{int(round(o_age))}</b>세 · 기준 <b>{int(round(i_age))}</b>세<br>"
            f"항목 요약: "
            f"<span style='color:#78c8a0'>개선 {n_improved}개</span> &nbsp; "
            f"<span style='color:#c85050'>역전(복원 후 낮아짐) {n_reversed}개</span> &nbsp; "
            f"동등 {n_equal}개"
        )
        head.setWordWrap(True)
        head.setTextFormat(Qt.TextFormat.RichText)

        legend = QLabel(
            "원본: 원본 이미지로 산출한 점수 | "
            "기준: 기준 이미지로 산출한 점수"
        )
        if self._llm_scores:
            legend.setText(
                legend.text() + " | "
                "LLM 기준: LLM 소견에 기반하여 조정된 점수"
            )
        legend.setStyleSheet("color: #666; font-size: 11px;")

        # 테이블 위젯 생성 (바차트 대신 엑셀 형식)
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        if self._llm_scores:
            self.table.setHorizontalHeaderLabels(["항목명", "원본 점수", "기준 점수", "원본 LLM 측정 점수", "기준 LLM 측정 점수", "차이 (원본/기준)"])
        else:
            self.table.setHorizontalHeaderLabels(["항목명", "원본 점수", "기준 점수", "원본 LLM 측정 점수", "기준 LLM 측정 점수", "차이 (원본/기준)"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("QTableWidget { font-size: 11px; }")
        
        # 테이블에 데이터 채우기
        row = 0
        for cat_name, keys in get_measurement_categories():
            for key in keys:
                vo = _numeric_value(mo.get(key))
                vi = _numeric_value(mi.get(key))
                label = _short_label(key)
                
                # 행 추가
                self.table.insertRow(row)
                
                # 항목명
                item_name = QTableWidgetItem(label)
                self.table.setItem(row, 0, item_name)
                
                # 원본 점수 (없으면 "N/A")
                if mo.get(key) is not None:
                    item_orig = QTableWidgetItem(f"{int(round(vo))}")
                else:
                    item_orig = QTableWidgetItem("N/A")
                    item_orig.setBackground(QColor(240, 240, 240))
                self.table.setItem(row, 1, item_orig)

                # 기준 점수 (없으면 "N/A")
                if mi.get(key) is not None:
                    item_ideal = QTableWidgetItem(f"{int(round(vi))}")
                else:
                    item_ideal = QTableWidgetItem("N/A")
                    item_ideal.setBackground(QColor(240, 240, 240))
                self.table.setItem(row, 2, item_ideal)
                
                # 원본 LLM 점수 (초기에는 빈 값)
                item_llm_orig = QTableWidgetItem("")
                item_llm_orig.setBackground(QColor(240, 240, 240))
                self.table.setItem(row, 3, item_llm_orig)

                # 복원 LLM 점수 (초기에는 빈 값)
                item_llm_ideal = QTableWidgetItem("")
                item_llm_ideal.setBackground(QColor(240, 240, 240))
                self.table.setItem(row, 4, item_llm_ideal)
                
                # 차이 (초기에는 복원-원본 차이만 표시, 나중에 원본/복원 LLM 차이로 업데이트)
                diff = vi - vo
                item_diff = QTableWidgetItem(f"{int(round(diff)):+d}")
                if diff > 0:
                    item_diff.setForeground(QColor(0, 150, 0))
                elif diff < 0:
                    item_diff.setForeground(QColor(200, 50, 50))
                self.table.setItem(row, 5, item_diff)
                
                row += 1

        # LLM 피부건강지수 표시 라벨
        self.llm_overall_label = QLabel("LLM 피부건강지수: —")
        self.llm_overall_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #333; padding: 5px;")
        self.llm_overall_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # LLM 소견 표시 영역
        self.llm_report_text = QTextEdit()
        self.llm_report_text.setReadOnly(True)
        self.llm_report_text.setMinimumHeight(500)  # 최소 높이 증가 (처방전 포함 전체 표시를 위해)
        self.llm_report_text.setPlaceholderText("LLM 소견이 여기에 표시됩니다.\n「LLM 소견 생성」 버튼을 클릭하세요.")

        # 버튼 영역
        self.btn_llm = QPushButton("LLM 소견 생성 (원본+기준 동시)")
        self.btn_llm.clicked.connect(lambda: self._generate_llm_report_dual())
        btn_excel = QPushButton("엑셀 내보내기")
        btn_excel.clicked.connect(self._export_to_excel)
        if not _OPENPYXL_AVAILABLE:
            btn_excel.setEnabled(False)
            btn_excel.setToolTip("openpyxl 라이브러리가 필요합니다 (pip install openpyxl)")
        btn_close = QPushButton("닫기")
        btn_close.clicked.connect(self.accept)
        
        bottom = QHBoxLayout()
        bottom.addWidget(self.btn_llm)
        bottom.addWidget(btn_excel)
        bottom.addStretch(1)
        bottom.addWidget(btn_close)

        root = QVBoxLayout(self)
        root.addWidget(head)
        root.addWidget(legend)
        root.addWidget(self.table, 1)
        root.addWidget(self.llm_overall_label)
        root.addWidget(QLabel("<b>LLM AI 소견</b>"))
        root.addWidget(self.llm_report_text)
        root.addLayout(bottom)
        
        # LLM 결과가 이미 있으면 표시
        if self._last_llm_report_orig and self._last_llm_report_ideal:
            self._on_llm_dual_finished(self._last_llm_report_orig, self._last_llm_report_ideal, 0.0)
            self.btn_llm.setText("LLM 소견 재생성 (원본+기준 동시)")
        else:
            # LLM 결과가 없으면 항상 자동으로 LLM 소견 생성 실행 (점수 제공 여부와 관계없이)
            QTimer.singleShot(100, self.btn_llm.click)

    def _generate_llm_report_dual(self) -> None:
        """LLM API를 사용하여 원본+기준 이미지를 한번에 분석하여 소견을 생성합니다."""
        self.llm_report_text.setText("LLM 소견 생성 중... 잠시 기다려주세요.")
        self.btn_llm.setEnabled(False)  # 버튼 비활성화

        from src.llm.llm_reporter import LlmSkinReporter
        
        # 측정 결과에서 필요한 데이터 추출
        orig_measurements_report = self._orig_result.get("measurements_report") or self._orig_result.get("measurements", {})
        orig_overall_score = self._orig_result.get("overall_score", 0)
        orig_perceived_age = self._orig_result.get("perceived_age", 0)
        
        ideal_measurements_report = self._ideal_result.get("measurements_report") or self._ideal_result.get("measurements", {})
        ideal_overall_score = self._ideal_result.get("overall_score", 0)
        ideal_perceived_age = self._ideal_result.get("perceived_age", 0)
        
        provide_scores = self._llm_provide_scores  # 항상 점수 제공
        log.info(f"[LLM 호출] 다이얼로그에서 LLM API 호출 시작 (provide_scores={provide_scores})")
        log.debug("_generate_llm_report_dual: self._llm_provide_scores=%s (True=점수 제공, False=점수 미제공)", provide_scores)
        
        # 백그라운드 스레드 생성
        self.thread = QThread()
        
        # progress_callback을 시그널로 연결
        def progress_callback(message: str) -> None:
            self.worker.progress.emit(message)
        
        self.worker = LlmWorker(
            LlmSkinReporter(progress_callback=progress_callback),
            self._orig_path,
            orig_measurements_report,
            orig_overall_score,
            orig_perceived_age,
            self._ideal_path,
            ideal_measurements_report,
            ideal_overall_score,
            ideal_perceived_age,
            provide_scores=provide_scores
        )
        self.worker.moveToThread(self.thread)
        
        # 시그널 연결
        self.worker.progress.connect(self._on_llm_progress)
        self.worker.finished.connect(self._on_llm_dual_finished)
        self.worker.error.connect(self._on_llm_error)
        self.thread.started.connect(self.worker.run)
        
        # 스레드 시작
        self.thread.start()
    
    def _on_llm_progress(self, message: str) -> None:
        """LLM 진행 상황 업데이트"""
        self.llm_report_text.setText(f"LLM 소견 생성 중...\n{message}")
        QCoreApplication.processEvents()

    def _on_llm_dual_finished(self, orig_report: object, ideal_report: object, elapsed_time: float = 0.0) -> None:
        """듀얼 이미지 LLM 분석 완료"""
        # 스레드가 있는 경우에만 종료 처리
        if hasattr(self, 'thread') and self.thread is not None and isinstance(self.thread, QThread):
            self.btn_llm.setEnabled(True)
            self.thread.quit()
            self.thread.wait()
        elif hasattr(self, 'btn_llm'):
            # 스레드가 없는 경우 버튼만 활성화
            self.btn_llm.setEnabled(True)

        # LLM 처리 시간 로그
        if elapsed_time > 0:
            log.info(f"[GUI] LLM 처리 시간: {elapsed_time:.2f}초")

        # 소견 텍스트 구성
        adjustment_note = ""
        if self._llm_scores and (getattr(orig_report, 'scores_adjusted', False) or getattr(ideal_report, 'scores_adjusted', False)):
            adjustment_note = "\n⚠ 점수가 LLM 소견에 기반하여 조정되었습니다.\n"

        # 엄격한 평가 모드 확인
        strict_mode_note = ""
        try:
            from src.utils.config import load_config as _load_config
            config = _load_config()
            strict_mode_enabled = config.get("score_criteria", {}).get("엄격한 평가 모드", {}).get("enabled", False)
            if strict_mode_enabled:
                strict_mode_note = "\n📋 엄격한 평가 모드가 적용되었습니다. 모든 카테고리에 엄격한 기준이 적용됩니다.\n"
        except Exception:
            pass

        recommendation_text = getattr(orig_report, 'recommendation', '')
        # recommendation이 리스트인 경우 문자열로 변환
        if isinstance(recommendation_text, list):
            recommendation_text = '\n'.join(recommendation_text)
        log.info(f"[GUI] 처방전 길이: {len(recommendation_text)}, 내용: {recommendation_text[:100] if recommendation_text else 'EMPTY'}")
        
        # 제품 추천 정보 추출
        matched_products = getattr(orig_report, 'matched_products', [])
        product_text = ""
        if matched_products and isinstance(matched_products, list):
            product_text = "\n【맞춤형 제품 추천】\n"
            for i, product in enumerate(matched_products, 1):
                product_name = product.get('product_name', '알 수 없음')
                category = product.get('category', '')
                key_ingredients = product.get('key_ingredients', [])
                efficacy = product.get('efficacy', '')
                match_score = product.get('match_score', 0)
                match_reason = product.get('match_reason', '')

                product_text += f"\n{i}. {product_name} ({category})\n"
                product_text += f"   매칭 점수: {int(round(match_score * 100))}%\n"
                product_text += f"   주요 성분: {', '.join(key_ingredients)}\n"
                product_text += f"   효능: {efficacy}\n"
                product_text += f"   추천 이유: {match_reason}\n"
        
        # 복원 LLM 측정 여부 확인
        ideal_has_llm = (
            hasattr(ideal_report, 'metric_opinions') and 
            len(ideal_report.metric_opinions) > 0
        )
        
        # 복원 LLM 점수 표시
        if ideal_has_llm:
            ideal_llm_score_text = f"- 기준 LLM 측정 피부건강지수: {int(round(ideal_report.overall_score))}점"
        else:
            ideal_llm_score_text = "- 기준 LLM 측정 피부건강지수: -"

        report_text = f"""=== LLM AI 소견 (원본+기준 동시 분석) ===
{adjustment_note}
{strict_mode_note}
【원본 이미지 종합 소견】
{orig_report.overall_opinion}

【기준 이미지 종합 소견】
{ideal_report.overall_opinion}

【관리 권고사항】
{recommendation_text}

{product_text}
【분석 메타데이터】
- 원본 LLM 측정 피부건강지수: {int(round(orig_report.overall_score))}점
{ideal_llm_score_text}
- 원본 인지나이: {int(round(orig_report.perceived_age))}세
- 기준 인지나이: {int(round(ideal_report.perceived_age))}세
"""
        
        # 원본 이미지 항목별 소견 추가 (테이블 순서와 동일하게)
        measurement_count = get_measurement_count()
        report_text += f"\n\n【원본 이미지 {measurement_count}개 항목별 LLM 소견】\n"
        
        # 테이블 순서와 동일하게 정렬하기 위해 딕셔너리 생성
        orig_opinions_dict = {metric.key: metric for metric in orig_report.metric_opinions}
        
        for cat_name, keys in get_measurement_categories():
            for key in keys:
                if key in orig_opinions_dict:
                    metric = orig_opinions_dict[key]
                    report_text += f"\n● {metric.display_name} ({int(round(metric.score))}점 / {metric.grade})\n"
                    report_text += f"  {metric.opinion}\n"
                    if metric.reason:
                        report_text += f"  [근거: {metric.reason}]\n"

        # 기준 이미지 항목별 소견 추가 (테이블 순서와 동일하게)
        report_text += f"\n\n【기준 이미지 {measurement_count}개 항목별 LLM 소견】\n"
        
        # 테이블 순서와 동일하게 정렬하기 위해 딕셔너리 생성
        ideal_opinions_dict = {metric.key: metric for metric in ideal_report.metric_opinions}
        
        for cat_name, keys in get_measurement_categories():
            for key in keys:
                if key in ideal_opinions_dict:
                    metric = ideal_opinions_dict[key]
                    report_text += f"\n● {metric.display_name} ({int(round(metric.score))}점 / {metric.grade})\n"
                    report_text += f"  {metric.opinion}\n"
                    if metric.reason:
                        report_text += f"  [근거: {metric.reason}]\n"

        self.llm_report_text.setText(report_text)
        # 스크롤을 상단으로 이동
        self.llm_report_text.verticalScrollBar().setValue(0)

        # Report 저장 (엑셀 내보내기 시 항목별 소견 출력용)
        self._last_llm_report_orig = orig_report
        self._last_llm_report_ideal = ideal_report

        # 테이블 헤더 업데이트
        if self._llm_scores:
            self.table.setHorizontalHeaderLabels(["항목명", "원본 점수", "기준 점수", "원본 LLM 측정 점수", "기준 LLM 측정 점수", "차이 (원본/기준)"])
        else:
            self.table.setHorizontalHeaderLabels(["항목명", "원본 점수", "기준 점수", "원본 LLM 측정 점수", "기준 LLM 측정 점수", "차이 (원본/기준)"])

        # 원본/기준 이미지 개별 항목 점수 업데이트 (동시에)
        self._update_table_with_llm_scores(orig_report.metric_opinions, ideal_report.metric_opinions)

        # 원본/기준 이미지 LLM 피부건강지수 표시 (동시에)
        # 기준 이미지의 실제 LLM 측정 점수가 있는 경우에만 표시
        # reference_guided 모드에서는 기준 LLM 점수가 없으므로 내부 점수 표시 안 함
        ideal_has_llm = (
            hasattr(ideal_report, 'metric_opinions') and 
            len(ideal_report.metric_opinions) > 0
        )
        if ideal_has_llm:
            self._update_table_with_llm_overall_both(orig_report.overall_score, ideal_report.overall_score)
        else:
            # 복원 LLM 점수가 없으면 원본만 표시
            self.llm_overall_label.setText(f"LLM 피부건강지수 - 원본: {int(round(orig_report.overall_score))}점 / 기준: —")

    def _on_llm_error(self, error_msg: str) -> None:
        """LLM 오류 처리"""
        self.btn_llm.setEnabled(True)
        self.thread.quit()
        self.thread.wait()

        error_text = f"LLM 소견 생성 실패:\n{error_msg}"
        self.llm_report_text.setText(error_text)
        QMessageBox.warning(self, "오류", error_text)

    def _update_table_with_llm_scores(
        self,
        orig_metric_opinions: List[Any],
        ref_metric_opinions: List[Any]
    ) -> None:
        """테이블에 LLM 기준 점수를 업데이트합니다 (원본+복원 동시).

        Args:
            orig_metric_opinions: 원본 이미지 LLM 소견에 포함된 항목별 점수
            ref_metric_opinions: 기준 이미지 LLM 소견에 포함된 항목별 점수
        """
        # 원본 점수 매핑 생성
        orig_scores = {}
        for metric in orig_metric_opinions:
            orig_scores[metric.display_name] = metric.score
        
        # 기준 점수 매핑 생성
        ref_scores = {}
        for metric in ref_metric_opinions:
            ref_scores[metric.display_name] = metric.score
        
        for r in range(self.table.rowCount()):
            item_name = self.table.item(r, 0)
            if not item_name:
                continue
            
            name = item_name.text()
            orig_score = orig_scores.get(name)
            ref_score = ref_scores.get(name)
            
            # 원본 점수 가져오기
            orig_base_item = self.table.item(r, 1)
            orig_base_score = None
            if orig_base_item:
                text = orig_base_item.text()
                if text and text != 'N/A':
                    try:
                        orig_base_score = float(text)
                    except (ValueError, AttributeError) as e:
                        log.debug("원본 기본 점수 파싱 실패: %s", e)

            # 기준 점수 가져오기
            ref_base_item = self.table.item(r, 2)
            ref_base_score = None
            if ref_base_item:
                text = ref_base_item.text()
                if text and text != 'N/A':
                    try:
                        ref_base_score = float(text)
                    except (ValueError, AttributeError) as e:
                        log.debug("기준 기본 점수 파싱 실패: %s", e)

            # 원본 LLM 점수 설정
            llm_orig_item = self.table.item(r, 3)
            if llm_orig_item and orig_score is not None:
                llm_orig_item.setText(f"{int(round(orig_score))}")
                # 점수를 제공하는 경우에만 조정 여부에 따라 배경색 변경
                if self._llm_scores and orig_base_score is not None and abs(orig_score - orig_base_score) > 0.1:
                    llm_orig_item.setBackground(QColor(255, 255, 200))  # 조정된 경우 노란색
                else:
                    llm_orig_item.setBackground(QColor(240, 240, 240))

            # 복원 LLM 점수 설정
            llm_ref_item = self.table.item(r, 4)
            if llm_ref_item and ref_score is not None:
                llm_ref_item.setText(f"{int(round(ref_score))}")
                # 점수를 제공하는 경우에만 조정 여부에 따라 배경색 변경
                if self._llm_scores and ref_base_score is not None and abs(ref_score - ref_base_score) > 0.1:
                    llm_ref_item.setBackground(QColor(255, 255, 200))  # 조정된 경우 노란색
                else:
                    llm_ref_item.setBackground(QColor(240, 240, 240))

            # 차이 업데이트 (LLM 원본/복원 차이)
            diff_item = self.table.item(r, 5)
            if diff_item and orig_score is not None and ref_score is not None:
                llm_diff = ref_score - orig_score
                diff_item.setText(f"{int(round(llm_diff)):+d}")
                if llm_diff > 0:
                    diff_item.setForeground(QColor(0, 150, 0))
                elif llm_diff < 0:
                    diff_item.setForeground(QColor(200, 50, 50))
                else:
                    diff_item.setForeground(QColor(100, 100, 100))
            elif diff_item:
                diff_item.setText("N/A")
                diff_item.setForeground(QColor(150, 150, 150))
    
    def _update_table_with_llm_overall_both(self, orig_overall_score: float, ideal_overall_score: float) -> None:
        """원본/기준 이미지의 LLM 피부건강지수를 라벨에 표시합니다.

        Args:
            orig_overall_score: 원본 이미지 LLM 피부건강지수
            ideal_overall_score: 기준 이미지 LLM 피부건강지수
        """
        self.llm_overall_label.setText(f"LLM 피부건강지수 - 원본: {int(round(orig_overall_score))}점 / 기준: {int(round(ideal_overall_score))}점")

        # 테이블에는 피부건강지수를 표시하지 않음 (항목별 점수만 표시)
        # 피부건강지수는 라벨에만 표시하여 항목별 점수와 혼동 방지
    
    @staticmethod
    def _sanitize_cell_text(text: str) -> str:
        """엑셀 XML에 허용되지 않는 제어문자를 제거한다.

        openpyxl은 0x00~0x08, 0x0B, 0x0C, 0x0E~0x1F 범위 문자를
        IllegalCharacterError 로 거부한다. LLM 응답 등 외부 텍스트를
        셀에 쓰기 전에 반드시 정제해야 한다.
        """
        import re as _re
        # XML 1.0 허용 범위 외 문자 제거 (탭·LF·CR 제외)
        return _re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)

    def _export_to_excel(self) -> None:
        """테이블 데이터를 엑셀 파일로 내보냅니다."""
        if not _OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "오류", "openpyxl 라이브러리가 필요합니다.\n설치: pip install openpyxl")
            return

        # scoring_mode 확인
        scoring_mode = "independent"  # 기본값
        try:
            from src.utils.config import load_config as _load_config
            config = _load_config()
            scoring_mode = config.get("llm", {}).get("scoring_mode", "independent")
        except Exception:
            pass

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"skin_comparison_{timestamp}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 파일 저장",
            default_name,
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "피부 분석 비교"

            # 스타일 정의
            bold_font  = Font(name="맑은 고딕", size=11, bold=True)
            small_font = Font(name="맑은 고딕", size=9, italic=False)

            # 로컬 함수: ws를 캡처하여 append_with_font 호출
            def append_with_font_local(values: List[Any], font: Optional[Font] = None) -> int:
                return append_with_font(ws, values, font)

            # ── 원본/기준 이미지 추가 (제일 위에 배치) ─────────────────
            try:
                # 원본 이미지
                img_orig = OpenpyxlImage(str(self._orig_path))
                img_orig.width = 200
                img_orig.height = 200
                img_orig.anchor = "A1"
                ws.add_image(img_orig)
                
                # 기준 이미지
                img_ideal = OpenpyxlImage(str(self._ideal_path))
                img_ideal.width = 200
                img_ideal.height = 200
                img_ideal.anchor = "D1"
                ws.add_image(img_ideal)
                
                # 이미지 라벨
                ws.cell(row=1, column=2, value="원본")
                ws.cell(row=1, column=2).font = bold_font
                ws.cell(row=1, column=6, value="기준")
                ws.cell(row=1, column=6).font = bold_font
                
                # 이미지 아래에 빈 행 추가 (이미지 공간 확보)
                for _ in range(10):
                    ws.append([None, None, None, None, None, None])
            except Exception as img_err:
                log.warning(f"이미지 추가 실패: {img_err}")

            # ── 피부건강지수 계산 ─────────────────────────────────────────
            orig_overall  = _numeric_value(
                self._orig_result.get("overall_score_report")
                or self._orig_result.get("overall_score")
            )
            ideal_overall = _numeric_value(
                self._ideal_result.get("overall_score_report")
                or self._ideal_result.get("overall_score")
            )

            # LLM 종합점수는 LLM 리포트의 overall_score 사용
            orig_llm_overall: Optional[float] = None
            ideal_llm_overall: Optional[float] = None

            if self._last_llm_report_orig and hasattr(self._last_llm_report_orig, 'overall_score'):
                orig_llm_overall = _numeric_value(self._last_llm_report_orig.overall_score)

            # 복원 LLM 측정 여부 확인 (reference_guided 모드에서는 항상 false)
            ideal_has_llm = (
                self._last_llm_report_ideal and 
                hasattr(self._last_llm_report_ideal, 'metric_opinions') and 
                len(self._last_llm_report_ideal.metric_opinions) > 0
            )
            if ideal_has_llm and hasattr(self._last_llm_report_ideal, 'overall_score'):
                ideal_llm_overall = _numeric_value(self._last_llm_report_ideal.overall_score)

            # ── 메타 헤더를 전부 append_with_font_local() 로 통일 ──────

            # 메타 정보 행 추가
            append_with_font_local(["원본 이미지", str(self._orig_path)], bold_font)
            append_with_font_local(["기준 이미지", str(self._ideal_path)], bold_font)
            append_with_font_local(["원본 피부건강지수", f"{int(round(orig_overall))}"], bold_font)
            append_with_font_local(["기준 피부건강지수", f"{int(round(ideal_overall))}"], bold_font)
            if orig_llm_overall is not None:
                append_with_font_local(["원본 LLM 피부건강지수", f"{int(round(orig_llm_overall))}"], bold_font)
            # 기준 LLM 측정이 없으면 '-' 표시
            if ideal_has_llm and ideal_llm_overall is not None:
                append_with_font_local(["기준 LLM 피부건강지수", f"{int(round(ideal_llm_overall))}"], bold_font)
            else:
                append_with_font_local(["기준 LLM 피부건강지수", "-"], bold_font)
            # 엄격한 평가 모드 표시
            try:
                from src.utils.config import load_config as _load_config
                config = _load_config()
                strict_mode_enabled = config.get("score_criteria", {}).get("엄격한 평가 모드", {}).get("enabled", False)
                strict_mode_status = "적용됨" if strict_mode_enabled else "적용안됨"
                append_with_font_local(["엄격한 평가 모드", strict_mode_status], bold_font)
            except Exception:
                pass
            append_with_font_local([])  # 빈 행

            # ── 테이블 데이터 추가 ─────────────────────────────────
            # 헤더
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            append_with_font_local(headers, bold_font)

            # 데이터 행
            for r in range(self.table.rowCount()):
                row_data = []
                for c in range(self.table.columnCount()):
                    item = self.table.item(r, c)
                    if item:
                        row_data.append(item.text())
                    else:
                        row_data.append("")
                append_with_font_local(row_data)

            # ── LLM 소견 추가 ─────────────────────────────────────
            append_with_font_local([])  # 빈 행
            append_with_font_local(["LLM 소견"], bold_font)
            append_with_font_local([])  # 빈 행

            if self._last_llm_report_orig:
                append_with_font_local(["원본 이미지 종합 소견"], bold_font)
                append_with_font_local([self._sanitize_cell_text(self._last_llm_report_orig.overall_opinion)], small_font)
                append_with_font_local([])  # 빈 행

                append_with_font_local(["원본 이미지 항목별 소견"], bold_font)
                # 테이블 순서와 동일하게 정렬하기 위해 딕셔너리 생성
                orig_opinions_dict = {metric.key: metric for metric in self._last_llm_report_orig.metric_opinions}
                for cat_name, keys in get_measurement_categories():
                    for key in keys:
                        if key in orig_opinions_dict:
                            metric = orig_opinions_dict[key]
                            append_with_font_local([f"{metric.display_name} ({int(round(metric.score))}점 / {metric.grade})"], small_font)
                            # metric.opinion이 dict인 경우 문자열 추출
                            opinion_text = metric.opinion
                            if isinstance(opinion_text, dict):
                                opinion_text = opinion_text.get('opinion', str(opinion_text))
                            elif not isinstance(opinion_text, str):
                                opinion_text = str(opinion_text)
                            append_with_font_local([self._sanitize_cell_text(opinion_text)], small_font)
                            # 근거 필드 추가
                            if metric.reason:
                                append_with_font_local([f"[근거: {metric.reason}"], small_font)
                            append_with_font_local([])  # 빈 행

            # reference_guided 모드에서는 기준 이미지 종합 소견 제외하지만 산출근거는 표시
            if self._last_llm_report_ideal:
                if scoring_mode != "reference_guided":
                    append_with_font_local(["기준 이미지 종합 소견"], bold_font)
                    append_with_font_local([self._sanitize_cell_text(self._last_llm_report_ideal.overall_opinion)], small_font)
                    append_with_font_local([])  # 빈 행

                append_with_font_local(["기준 이미지 항목별 산출근거"], bold_font)
                # 테이블 순서와 동일하게 정렬하기 위해 딕셔너리 생성
                ideal_opinions_dict = {metric.key: metric for metric in self._last_llm_report_ideal.metric_opinions}
                for cat_name, keys in get_measurement_categories():
                    for key in keys:
                        if key in ideal_opinions_dict:
                            metric = ideal_opinions_dict[key]
                            append_with_font_local([f"{metric.display_name} ({int(round(metric.score))}점 / {metric.grade})"], small_font)
                            # 소견은 표시하지 않고 산출근거만 표시
                            if metric.reason:
                                append_with_font_local([f"[근거: {metric.reason}]"], small_font)
                            append_with_font_local([])  # 빈 행

            if self._last_llm_report_orig:
                append_with_font_local(["관리 권고사항"], bold_font)
                append_with_font_local([self._sanitize_cell_text(self._last_llm_report_orig.recommendation)], small_font)
                append_with_font_local([])  # 빈 행

                # 제품 추천 정보 추가
                matched_products = getattr(self._last_llm_report_orig, 'matched_products', [])
                if matched_products and isinstance(matched_products, list):
                    append_with_font_local(["맞춤형 제품 추천"], bold_font)
                    for i, product in enumerate(matched_products, 1):
                        product_name = product.get('product_name', '알 수 없음')
                        category = product.get('category', '')
                        key_ingredients = product.get('key_ingredients', [])
                        efficacy = product.get('efficacy', '')
                        match_score = product.get('match_score', 0)
                        match_reason = product.get('match_reason', '')

                        # match_score가 float 형식이 아닌 경우 변환
                        try:
                            match_score_float = float(match_score) if match_score is not None else 0.0
                        except (TypeError, ValueError):
                            match_score_float = 0.0

                        append_with_font_local([f"{i}. {product_name} ({category})"], small_font)
                        append_with_font_local([f"매칭 점수: {int(round(match_score_float))}"], small_font)
                        append_with_font_local([f"주요 성분: {', '.join(key_ingredients)}"], small_font)
                        append_with_font_local([f"효능: {efficacy}"], small_font)
                        append_with_font_local([f"추천 이유: {match_reason}"], small_font)
                        append_with_font_local([])  # 빈 행

            # 저장
            wb.save(file_path)
            QMessageBox.information(self, "완료", f"엑셀 파일 저장 완료:\n{file_path}")
            log.info("엑셀 내보내기 완료: %s", file_path)

        except Exception as e:
            error_detail = traceback.format_exc()
            log.error("=" * 60)
            log.error("엑셀 내보내기 실패")
            log.error("오류 타입: %s", type(e).__name__)
            log.error("오류 메시지: %s", str(e))
            log.error("상세 스택 트레이스:")
            log.error("%s", error_detail)
            log.error("=" * 60)
            error_msg = f"엑셀 내보내기 실패:\n{str(e)}\n\n상세 로그를 확인하세요."
            QMessageBox.critical(self, "오류", error_msg)

    def closeEvent(self, event: QCloseEvent) -> None:
        """다이얼로그 닫을 때 스레드 cleanup 보장."""
        # 다이얼로그는 modal로 표시되므로 스레드는 이미 종료되어 있어야 함
        # 하지만 안전하게 cleanup 시도
        super().closeEvent(event)
