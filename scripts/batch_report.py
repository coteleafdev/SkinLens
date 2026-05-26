"""
배치 보고서 생성 스크립트

images 폴더의 이미지들을 순회하며 파이프라인을 실행하고,
엑셀 파일에 원본이미지, 복원이미지, 점수를 횡배치로 표시하는 보고서를 생성합니다.
"""

import json
import os
import shutil
import sys
from pathlib import Path
from datetime import datetime

# 프로젝트 루트를 sys.path에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("오류: openpyxl 라이브러리가 필요합니다.")
    print("설치: pip install openpyxl")
    sys.exit(1)

from src.pipeline.pipeline_core import (
    Restorer,
    PipelineSettings,
    final_pipeline_artifact_path,
    format_duration,
    resolve_init_image,
    run_enhancement_pipeline,
)
from src.utils.utils import apply_score_safety_net


# ─────────────────────────────────────────────────────────────────────────────
#  모듈 레벨 유틸 함수
# ─────────────────────────────────────────────────────────────────────────────

def _safe_load_config() -> dict:
    """config.json 단일 로드 함수.

    [FIX DESIGN-3] get_restoration_defaults 와 create_report 루프 양쪽에서
    _load_scoring_config 를 각각 호출하던 중복을 제거.
    create_report 상단에서 한 번 호출하고 결과를 하위 함수에 전달한다.
    """
    try:
        from src.scoring.skin_scoring import _load_scoring_config
        return _load_scoring_config() or {}
    except Exception:
        return {}


def apply_score_offset(
    score_data: dict,
    offset_config: dict,
    weights: dict,
) -> dict:
    """세부 항목별 offset 배분 후 종합점수 재계산.

    [FIX BUG-4] 루프 내 중첩 정의 → 모듈 레벨 함수로 이동.
    [FIX BUG-5] _raw 접미사 키를 total_weight 합산에서 제외.
                _raw 키가 포함되면 weights 에 해당 키가 없어
                total_weight 가 작아지고 item_offset 이 과대 계산됨.
    """
    if not offset_config.get("enabled", False):
        return score_data

    offset = offset_config.get("offset", 0.0)
    if offset == 0.0:
        return score_data

    measurements = score_data.get("measurements", {})
    # [FIX BUG-5] _raw 키 제외 후 합산
    total_weight = sum(
        weights.get(k, 0.0)
        for k in measurements
        if not k.endswith("_raw")
    )

    adjusted_measurements = {}
    for key, value in measurements.items():
        if key.endswith("_raw"):          # _raw 키는 그대로 통과
            adjusted_measurements[key] = value
            continue
        weight = weights.get(key, 0.0)
        if total_weight > 0 and weight > 0:
            item_offset = offset * (weight / total_weight)
            adjusted_measurements[key] = min(90.0, float(value) + item_offset)
        else:
            adjusted_measurements[key] = value

    overall = score_data.get("overall", 0.0)
    adjusted_overall = min(90.0, float(overall) + offset)

    return {
        "overall": adjusted_overall,
        "measurements": adjusted_measurements,
    }


def get_restoration_defaults(scoring_config: dict | None = None):
    """config.json에서 복원 파라미터 기본값을 가져옵니다.

    [FIX DESIGN-3] scoring_config 를 외부에서 주입받아 중복 파일 I/O 방지.
    미제공 시 _safe_load_config() 로 로드.
    """
    config = scoring_config if scoring_config is not None else _safe_load_config()
    restoration = config.get("restoration", {})
    return (
        restoration.get("codeformer_fidelity", 1.0),
        restoration.get("codeformer_upscale", 1),
        restoration.get("codeformer_bg_upsampler", "none"),
    )


def setup_excel_styles(ws):
    """엑셀 스타일 설정"""
    # 헤더 스타일
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # 셀 스타일
    cell_font = Font(name='맑은 고딕', size=10)
    cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    return {
        'header_font': header_font,
        'header_fill': header_fill,
        'header_alignment': header_alignment,
        'header_border': header_border,
        'cell_font': cell_font,
        'cell_alignment': cell_alignment,
        'cell_border': cell_border
    }


def add_image_to_excel(ws, img_path, row, col, width=15, height=15):
    """엑셀 시트에 이미지 추가"""
    try:
        img = OpenpyxlImage(img_path)
        img.width = width * 7  # 픽셀로 변환 (대략적인 비율)
        img.height = height * 7
        img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img)
        return True
    except Exception as e:
        print(f"이미지 추가 실패: {img_path}, 오류: {e}")
        return False


def create_report(images_dir, output_excel, out_dir):
    """배치 보고서 생성"""
    images_path = Path(images_dir)
    output_path = Path(output_excel)
    out_path = Path(out_dir)

    # 보고서와 함께 저장할 이미지 폴더 생성
    report_images_dir = output_path.parent / f"{output_path.stem}_images"
    report_images_dir.mkdir(parents=True, exist_ok=True)

    if not images_path.exists():
        print(f"오류: 이미지 폴더가 존재하지 않습니다: {images_path}")
        return False

    # 이미지 파일 목록 가져오기 (모든 이미지 파일 처리)
    image_extensions = {'.jpg', '.jpeg', '.png', '.bmp'}
    image_files = [f for f in images_path.iterdir() if f.suffix.lower() in image_extensions]

    if not image_files:
        print(f"오류: 이미지 파일이 없습니다: {images_path}")
        return False

    print(f"총 {len(image_files)}개 이미지 발견")

    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "보고서"

    # 스타일 설정
    styles = setup_excel_styles(ws)

    # 헤더 행
    headers = ["번호", "파일명", "원본 이미지", "복원 이미지", "원본 점수", "복원 점수", "차이"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['header_border']

    # 열 너비 설정
    ws.column_dimensions['A'].width = 5  # 번호
    ws.column_dimensions['B'].width = 25  # 파일명
    ws.column_dimensions['C'].width = 15  # 원본 이미지
    ws.column_dimensions['D'].width = 15  # 복원 이미지
    ws.column_dimensions['E'].width = 35  # 원본 점수
    ws.column_dimensions['F'].width = 35  # 복원 점수
    ws.column_dimensions['G'].width = 10  # 차이

    # [FIX DESIGN-3] scoring_config 단일 로드 — 루프 내 반복 로드 제거
    scoring_config = _safe_load_config()
    offset_config  = scoring_config.get("score_offset", {})
    
    # 측정항목 메타데이터는 config_parser에서 로드 (SSOT 원칙)
    from src.scoring.skin_scoring import get_measurement_weights, get_display_names
    weights        = get_measurement_weights()
    display_names  = get_display_names()

    # 복원 파라미터 (로드된 config 재사용)
    codeformer_fidelity, codeformer_upscale, codeformer_bg_upsampler = get_restoration_defaults(scoring_config)
    print(f"복원 파라미터: fidelity={codeformer_fidelity}, upscale={codeformer_upscale}, bg_upsampler={codeformer_bg_upsampler}")

    # 파이프라인 설정
    cfg = PipelineSettings()
    cfg.restorer = Restorer.CODEFORMER
    cfg.codeformer_fidelity = codeformer_fidelity
    cfg.codeformer_upscale = codeformer_upscale
    cfg.codeformer_bg_upsampler = codeformer_bg_upsampler

    # 각 이미지 처리
    for idx, img_file in enumerate(image_files, 1):
        print(f"\n[{idx}/{len(image_files)}] 처리 중: {img_file.name}")

        try:
            # 파이프라인 실행
            print(f"  파이프라인 실행 중...")
            r = run_enhancement_pipeline(
                cfg=cfg,
                out_dir=out_path,
                input_image=str(img_file),
                do_restore=True,
            )

            # 최종 산출물 경로
            final_p = final_pipeline_artifact_path(r, out_path)
            if final_p is None or not final_p.is_file():
                print(f"  오류: 최종 산출물 없음")
                continue

            # 1024x1024 리사이즈된 원본 이미지 복사
            stem = r.output_stem
            input_resized = out_path / f"00_input_{stem}.png"
            if input_resized.is_file():
                orig_copy = report_images_dir / f"{idx:03d}_original_{img_file.stem}.png"
                shutil.copy2(input_resized, orig_copy)
                print(f"  원본 이미지 복사: {orig_copy.name}")

            # 복원 이미지 복사
            restored_copy = report_images_dir / f"{idx:03d}_restored_{img_file.stem}.png"
            shutil.copy2(final_p, restored_copy)
            print(f"  복원 이미지 복사: {restored_copy.name}")

            # 점수 분석 및 안전장치 적용
            print(f"  점수 분석 중...")
            try:
                o, i1, i2 = apply_score_safety_net(img_file, final_p)
            except Exception as e:
                print(f"  경고: 점수 안전장치 실패: {e}")
                from src.skin.core.analyze_utils import analyze_compare_triple
                o, i1, i2 = analyze_compare_triple(img_file, final_p, final_p)

            # 점수 추출
            orig_overall = float(o.get("overall_score_report", o.get("overall_score", 0)))
            restored_overall = float(i1.get("overall_score_report", i1.get("overall_score", 0)))
            diff = restored_overall - orig_overall

            # score_offset 적용
            # [FIX BUG-4] apply_score_offset은 모듈 레벨 함수로 이동됨

            # [FIX DESIGN-3] scoring_config 는 create_report 상단에서 단일 로드됨

            # 원본 점수 데이터
            orig_measurements = o.get("measurements_report") or o.get("measurements", {})
            orig_measurements_filtered = {k: v for k, v in orig_measurements.items() if not k.endswith("_raw")}
            orig_score_data = {
                "overall": orig_overall,
                "measurements": orig_measurements_filtered
            }

            # 복원 점수 데이터
            restored_measurements = i1.get("measurements_report") or i1.get("measurements", {})
            restored_measurements_filtered = {k: v for k, v in restored_measurements.items() if not k.endswith("_raw")}
            restored_score_data = {
                "overall": restored_overall,
                "measurements": restored_measurements_filtered
            }

            # offset 적용
            orig_score_adjusted = apply_score_offset(orig_score_data, offset_config, weights)
            restored_score_adjusted = apply_score_offset(restored_score_data, offset_config, weights)

            # 조정된 점수 사용
            orig_overall = orig_score_adjusted["overall"]
            restored_overall = restored_score_adjusted["overall"]
            orig_measurements_filtered = orig_score_adjusted["measurements"]
            restored_measurements_filtered = restored_score_adjusted["measurements"]
            diff = restored_overall - orig_overall

            # 엑셀에 데이터 추가
            row_idx = idx + 1

            # 번호
            cell = ws.cell(row=row_idx, column=1, value=idx)
            cell.font = styles['cell_font']
            cell.alignment = styles['cell_alignment']
            cell.border = styles['cell_border']

            # 파일명
            cell = ws.cell(row=row_idx, column=2, value=img_file.name)
            cell.font = styles['cell_font']
            cell.alignment = styles['cell_alignment']
            cell.border = styles['cell_border']

            # 원본 이미지
            add_image_to_excel(ws, img_file, row_idx, 3, width=15, height=15)
            ws.row_dimensions[row_idx].height = 350  # 18개 측정항목이 한눈에 보이도록 행 높이 조정

            # 복원 이미지
            add_image_to_excel(ws, final_p, row_idx, 4, width=15, height=15)

            # 원본 점수
            orig_score_text = f"종합: {orig_overall:.1f}\n"
            for key, val in sorted(orig_measurements_filtered.items()):
                if key != "overall_score_report":
                    korean_name = display_names.get(key, key)
                    orig_score_text += f"{korean_name}: {val:.1f}\n"

            cell = ws.cell(row=row_idx, column=5, value=orig_score_text)
            cell.font = styles['cell_font']
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = styles['cell_border']

            # 복원 점수
            restored_score_text = f"종합: {restored_overall:.1f}\n"
            for key, val in sorted(restored_measurements_filtered.items()):
                if key != "overall_score_report":
                    korean_name = display_names.get(key, key)
                    restored_score_text += f"{korean_name}: {val:.1f}\n"

            cell = ws.cell(row=row_idx, column=6, value=restored_score_text)
            cell.font = styles['cell_font']
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = styles['cell_border']

            # 차이
            diff_text = f"{diff:+.1f}"
            cell = ws.cell(row=row_idx, column=7, value=diff_text)
            cell.font = styles['cell_font']
            cell.alignment = styles['cell_alignment']
            cell.border = styles['cell_border']
            if diff >= 0:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # [FIX DESIGN-4]

            print(f"  완료: 원본 {orig_overall:.1f} → 복원 {restored_overall:.1f} ({diff:+.1f})")

            # [OPT-3] 10건마다 임시 저장 — 중간 실패 시 완료된 행 보존
            if idx % 10 == 0:
                tmp_path = output_path.with_suffix(".tmp.xlsx")
                try:
                    wb.save(tmp_path)
                    print(f"  [임시 저장] {tmp_path.name}")
                except Exception as _tmp_e:
                    print(f"  [임시 저장 실패] {_tmp_e}")

        except Exception as e:
            print(f"  오류: {e}")
            import traceback
            traceback.print_exc()
            continue

    # 엑셀 파일 저장
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n보고서 저장 완료: {output_path}")
    print(f"이미지 저장 완료: {report_images_dir}")
    return True


def main():
    """메인 함수"""
    # 기본 설정
    images_dir = "images"
    output_excel = f"batch_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_dir = "results"

    # 사용자 정의 가능 (명령행 인자)
    if len(sys.argv) > 1:
        images_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_excel = sys.argv[2]
    if len(sys.argv) > 3:
        out_dir = sys.argv[3]

    print("=" * 60)
    print("배치 보고서 생성")
    print("=" * 60)
    print(f"이미지 폴더: {images_dir}")
    print(f"출력 엑셀: {output_excel}")
    print(f"출력 폴더: {out_dir}")
    print("=" * 60)

    # 보고서 생성
    success = create_report(images_dir, output_excel, out_dir)

    if success:
        print("\n배치 보고서 생성 완료!")
        return 0
    else:
        print("\n배치 보고서 생성 실패!")
        return 1


if __name__ == "__main__":
    sys.exit(main())
